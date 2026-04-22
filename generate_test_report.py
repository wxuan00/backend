#!/usr/bin/env python3
"""
generate_test_report.py
Reads Maven Surefire XML reports from target/surefire-reports/
and exports a formatted Excel workbook: test-report.xlsx
"""

import os, sys, glob, xml.etree.ElementTree as ET
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SUREFIRE_DIR = os.path.join(os.path.dirname(__file__), "target", "surefire-reports")
OUTPUT_FILE  = os.path.join(os.path.dirname(__file__), "test-report.xlsx")
PROJECT_NAME = "Merchant Service Portal (MSP) - Backend Unit Test Report"
GENERATED_AT = datetime.now().strftime("%Y-%m-%d %H:%M")

CLR_NAVY="1F3864"; CLR_PASS="C6EFCE"; CLR_FAIL="FFC7CE"; CLR_SKIP="FFEB9C"
CLR_SUMMARY="D9E1F2"; CLR_ALT="F2F2F2"; CLR_WHITE="FFFFFF"; CLR_BLUE="2F5496"

def thin_border():
    s=Side(border_style="thin",color="BFBFBF")
    return Border(left=s,right=s,top=s,bottom=s)

def sc(ws,row,col,value,bold=False,fill=None,halign="left",fc="000000",size=11,wrap=False):
    c=ws.cell(row=row,column=col,value=value)
    c.font=Font(name="Calibri",bold=bold,size=size,color=fc)
    c.alignment=Alignment(horizontal=halign,vertical="center",wrap_text=wrap)
    c.border=thin_border()
    if fill: c.fill=PatternFill("solid",fgColor=fill)
    return c

def hc(ws,row,col,value,size=11):
    return sc(ws,row,col,value,bold=True,fill=CLR_NAVY,halign="center",fc=CLR_WHITE,size=size)

def set_widths(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

def parse_reports(d):
    suites,cases=[],[]
    for f in sorted(glob.glob(os.path.join(d,"TEST-*.xml"))):
        try: root=ET.parse(f).getroot()
        except: continue
        name=root.attrib.get("name",os.path.basename(f))
        tot=int(root.attrib.get("tests",0))
        fail=int(root.attrib.get("failures",0))
        err=int(root.attrib.get("errors",0))
        skip=int(root.attrib.get("skipped",0))
        t=float(root.attrib.get("time",0))
        suites.append({"name":name,"total":tot,"passed":tot-fail-err-skip,
                        "failures":fail,"errors":err,"skipped":skip,"time":round(t,3)})
        for tc in root.findall("testcase"):
            cn=tc.attrib.get("classname",name); mn=tc.attrib.get("name","")
            ft=tc.find("failure"); et=tc.find("error"); sk=tc.find("skipped")
            if ft is not None: st,msg="FAIL",(ft.attrib.get("message",ft.text or ""))[:300]
            elif et is not None: st,msg="ERROR",(et.attrib.get("message",et.text or ""))[:300]
            elif sk is not None: st,msg="SKIP","Skipped"
            else: st,msg="PASS",""
            parts=cn.split(".")
            mod="Unknown"
            if "modules" in parts:
                idx=parts.index("modules")
                if idx+1<len(parts): mod=parts[idx+1].capitalize()
            elif "auth" in cn.lower(): mod="Auth"
            cases.append({"suite":name,"module":mod,"class":cn.split(".")[-1],
                           "method":mn,"display":mn.replace("_"," ").strip(),
                           "status":st,"time":round(float(tc.attrib.get("time",0))*1000,1),"message":msg})
    return suites,cases

def build_summary(wb,suites,cases):
    ws=wb.create_sheet("Summary",0); ws.sheet_view.showGridLines=False
    ws.merge_cells("A1:G2"); c=ws["A1"]
    c.value=PROJECT_NAME; c.font=Font(name="Calibri",bold=True,size=16,color=CLR_WHITE)
    c.fill=PatternFill("solid",fgColor=CLR_NAVY); c.alignment=Alignment(horizontal="center",vertical="center")
    ws.merge_cells("A3:G3"); s=ws["A3"]
    s.value=f"Generated: {GENERATED_AT}   |   Test Suites: {len(suites)}"
    s.font=Font(name="Calibri",italic=True,size=11,color="444444")
    s.fill=PatternFill("solid",fgColor="EBF3FB"); s.alignment=Alignment(horizontal="center",vertical="center")

    tt=sum(x["total"] for x in suites); tp=sum(x["passed"] for x in suites)
    tf=sum(x["failures"]+x["errors"] for x in suites); tsk=sum(x["skipped"] for x in suites)
    ttime=round(sum(x["time"] for x in suites),2)
    pr=f"{tp/tt*100:.1f}%" if tt else "N/A"
    kpis=[("Total Tests",tt,"4472C4"),("Passed",tp,"70AD47"),("Failed",tf,"C00000"),
          ("Skipped",tsk,"FFC000"),("Pass Rate",pr,"4472C4"),("Runtime (s)",ttime,"4472C4")]
    ws.row_dimensions[5].height=22; ws.row_dimensions[6].height=32
    for i,(lbl,val,clr) in enumerate(kpis,1):
        c=ws.cell(row=5,column=i,value=lbl); c.font=Font(name="Calibri",bold=True,size=10,color=CLR_WHITE)
        c.fill=PatternFill("solid",fgColor=clr); c.alignment=Alignment(horizontal="center",vertical="center")
        c=ws.cell(row=6,column=i,value=val); c.font=Font(name="Calibri",bold=True,size=15)
        c.fill=PatternFill("solid",fgColor=CLR_SUMMARY); c.alignment=Alignment(horizontal="center",vertical="center")
        c.border=thin_border()

    ws.row_dimensions[8].height=22
    for col,h in enumerate(["#","Test Suite","Module","Total","Passed","Failed/Error","Skipped","Time (s)","Pass Rate"],1):
        hc(ws,8,col,h)

    def mod_from(name):
        for kw in ["auth","user","merchant","transaction","refund","settlement","creditadvice","analytics","role","jwt","totp"]:
            if kw in name.lower(): return kw.capitalize()
        return "-"

    for i,s in enumerate(sorted(suites,key=lambda x:x["name"]),1):
        r=8+i; ws.row_dimensions[r].height=18
        bg=CLR_ALT if i%2==0 else CLR_WHITE
        nm=s["name"].split(".")[-1] if "." in s["name"] else s["name"]
        sf=s["failures"]+s["errors"]; sr=f"{s['passed']/s['total']*100:.1f}%" if s["total"] else "N/A"
        for col,v in enumerate([i,nm,mod_from(s["name"]),s["total"],s["passed"],sf,s["skipped"],s["time"],sr],1):
            cell=sc(ws,r,col,v,halign="center" if col!=2 else "left",fill=bg)
            if col==6 and sf>0: cell.font=Font(name="Calibri",bold=True,color="C00000"); cell.fill=PatternFill("solid",fgColor=CLR_FAIL)

    set_widths(ws,[5,40,14,9,9,14,9,11,11]); ws.freeze_panes="A9"

def build_all_tests(wb,cases):
    ws=wb.create_sheet("All Tests"); ws.sheet_view.showGridLines=False
    ws.row_dimensions[1].height=22
    for col,h in enumerate(["#","Module","Test Class","Test Method / Description","Status","Duration (ms)"],1): hc(ws,1,col,h)
    sfill={"PASS":CLR_PASS,"FAIL":CLR_FAIL,"ERROR":CLR_FAIL,"SKIP":CLR_SKIP}
    for i,tc in enumerate(cases,1):
        r=i+1; ws.row_dimensions[r].height=16; bg=CLR_ALT if i%2==0 else CLR_WHITE
        for col,v in enumerate([i,tc["module"],tc["class"],tc["display"],tc["status"],tc["time"]],1):
            cell=sc(ws,r,col,v,halign="left" if col==4 else "center",fill=bg)
            if col==5:
                cell.fill=PatternFill("solid",fgColor=sfill.get(tc["status"],CLR_WHITE))
                cell.font=Font(name="Calibri",bold=True,size=11)
    set_widths(ws,[6,14,32,58,10,16]); ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:F{len(cases)+1}"

def build_failures(wb,cases):
    ws=wb.create_sheet("Failures & Errors"); ws.sheet_view.showGridLines=False
    fl=[tc for tc in cases if tc["status"] in ("FAIL","ERROR")]
    if not fl:
        ws.merge_cells("A1:E2"); c=ws["A1"]
        c.value="All tests passed - no failures or errors!"; c.font=Font(name="Calibri",bold=True,size=13,color="375623")
        c.fill=PatternFill("solid",fgColor=CLR_PASS); c.alignment=Alignment(horizontal="center",vertical="center")
        set_widths(ws,[10,20,35,55,60]); return
    ws.row_dimensions[1].height=22
    for col,h in enumerate(["#","Module","Test Class","Test Method","Error Message"],1): hc(ws,1,col,h)
    for i,tc in enumerate(fl,1):
        r=i+1; ws.row_dimensions[r].height=40
        for col,v in enumerate([i,tc["module"],tc["class"],tc["display"],tc["message"]],1):
            sc(ws,r,col,v,fill=CLR_FAIL,halign="left" if col>=4 else "center",wrap=(col==5))
    set_widths(ws,[6,14,32,40,70]); ws.freeze_panes="A2"

def build_by_module(wb,cases):
    ws=wb.create_sheet("By Module"); ws.sheet_view.showGridLines=False
    mods={}
    for tc in cases:
        m=tc["module"]; mods.setdefault(m,{"total":0,"pass":0,"fail":0,"skip":0,"classes":set()})
        mods[m]["total"]+=1; mods[m]["classes"].add(tc["class"])
        if tc["status"]=="PASS": mods[m]["pass"]+=1
        elif tc["status"] in ("FAIL","ERROR"): mods[m]["fail"]+=1
        else: mods[m]["skip"]+=1
    ws.row_dimensions[1].height=22
    for col,h in enumerate(["Module","Test Classes","Total","Passed","Failed","Skipped","Pass Rate","Status"],1): hc(ws,1,col,h)
    for i,(mod,d) in enumerate(sorted(mods.items()),1):
        r=i+1; ws.row_dimensions[r].height=18; bg=CLR_ALT if i%2==0 else CLR_WHITE
        rate=f"{d['pass']/d['total']*100:.1f}%" if d["total"] else "N/A"
        st="All Pass" if d["fail"]==0 else f"{d['fail']} Failed"
        stfill=CLR_PASS if d["fail"]==0 else CLR_FAIL
        cls=", ".join(sorted(d["classes"]))
        for col,v in enumerate([mod,cls,d["total"],d["pass"],d["fail"],d["skip"],rate,st],1):
            cell=sc(ws,r,col,v,halign="left" if col in (1,2,8) else "center",fill=bg,wrap=(col==2))
            if col==8: cell.fill=PatternFill("solid",fgColor=stfill); cell.font=Font(name="Calibri",bold=True)
            if col==5 and d["fail"]>0: cell.font=Font(name="Calibri",bold=True,color="C00000")
    set_widths(ws,[16,52,10,10,10,10,12,16]); ws.freeze_panes="A2"

def main():
    if not os.path.isdir(SUREFIRE_DIR):
        print(f"ERROR: Surefire directory not found: {SUREFIRE_DIR}"); sys.exit(1)
    print(f"Reading: {SUREFIRE_DIR}")
    suites,cases=parse_reports(SUREFIRE_DIR)
    if not suites: print("No XML reports found. Run tests first."); sys.exit(1)
    tt=sum(s["total"] for s in suites); tp=sum(s["passed"] for s in suites)
    tf=sum(s["failures"]+s["errors"] for s in suites)
    print(f"Suites: {len(suites)}  Tests: {tt}  Passed: {tp}  Failed: {tf}")
    wb=openpyxl.Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    build_summary(wb,suites,cases)
    build_all_tests(wb,cases)
    build_failures(wb,cases)
    build_by_module(wb,cases)
    wb.save(OUTPUT_FILE)
    print(f"Report saved: {OUTPUT_FILE}")
    print("Sheets: Summary | All Tests | Failures & Errors | By Module")

if __name__=="__main__": main()
